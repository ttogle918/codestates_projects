import os
import sqlite3
from openpyxl import load_workbook

DATABASE_PATH = os.path.join(os.getcwd(), 'house.db')

conn = sqlite3.connect(DATABASE_PATH)
cur = conn.cursor()

load_wb = load_workbook("namyangju_code.xlsx", data_only=True)

load_ws = load_wb['Sheet1']

row, col = 1, 1
data = load_ws.cell(row, col).value

insert_code = "INSERT INTO Dong_code (id, do_name, si_name, eup_name, lat, lot) VALUES (?, ?, ?, ?, ?, ?)"
while row < 18 :
  temp = []
  col = 1
  while col < 7 :
    temp.append(load_ws.cell(row, col).value)
    col = col+1

  cur.execute(insert_code, (temp[0], temp[1], temp[2], temp[3], temp[4], temp[5]))
  row = row + 1
  
conn.commit()